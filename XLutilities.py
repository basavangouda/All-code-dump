import openpyxl

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    return(sheet.max_row)

def getColCount(file,sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    return(sheet.max_column)

def ReadData(file,sheetName,rownum,colmnum):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    return sheet.cell(row=rownum, column=colmnum).value


def WriteData(file,sheetName,rownum,colmnum,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(sheetName)
    sheet=workbook.active
    sheet.cell(row=rownum,column=colmnum).value=data
    workbook.save(file)