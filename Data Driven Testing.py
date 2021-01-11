#Reading the data from excel file
from selenium import webdriver
'''import openpyxl
path= "C:\\Users\\Lenovo\\Desktop\\PythonPractice\\XYZ.xlsx"
workbook = openpyxl.load_workbook(path)
sheet = workbook.active

rows = sheet.max_row
cols = sheet.max_column
print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r ,column=c).value, end="    ")
    print()'''

#how to write data into excel
import openpyxl

path="C:\\Users\\Lenovo\\Desktop\\PythonPractice\\XYZ.xlsx"
workbook=openpyxl.load_workbook(path)
workbook.get_sheet_by_name("Sheet2")
sheet=workbook.active

for r in range(1,6):
    for c in range(1,4):
        sheet.cell(row=r, column=c).value="Welcome to Python"
workbook.save(path)




